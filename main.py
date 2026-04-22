"""Main CLI for pediatric/NICU genetic report extraction and REDCap compatibility.

Primary modes
-------------
1) extract        : report PDF/TXT -> linked single-report JSON
2) import-redcap  : flat REDCap raw export -> linked patient-case JSON
3) export-redcap  : linked JSON -> flat REDCap-compatible CSV/TSV/XLSX rows
4) print-redcap-columns : derive/export canonical column order from a REDCap data dictionary

`main.py` is the canonical entry point. `medace.py` can remain a tiny compatibility
shim but should not carry separate logic.
"""

from __future__ import annotations

import argparse
import csv
import importlib
import json
import logging
import os
import urllib.error
import urllib.request
from collections import OrderedDict, defaultdict
from enum import Enum
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence, Type

from pydantic import BaseModel, ValidationError

from schema import (
    AnalysisType,
    DiagnosisStudyPeriod,
    Dosage,
    FindingClassification,
    GeneticDiagnosis,
    GeneticDiagnosisGroup,
    GeneticFinding,
    GeneticReportExtraction,
    InheritancePattern,
    LabClassification,
    LegacyPhenotypeNote,
    LegacyRepeatRow,
    LinkedGeneticReport,
    LinkedGeneticsCase,
    NICUDiagnosisType,
    PatientDemographics,
    PatientPhenotypeNote,
    ReferenceGenome,
    ReportMetadata,
    SecondaryFindings,
    Segregation,
    Sex,
    TestInformation,
    TestInterpretation,
    TestInterpretationResult,
    TestLab,
    TestTimeframe,
    TestType,
    Zygosity,
    build_format_instructions,
    compose_name,
    get_case_model,
    get_extraction_model,
    normalize_date,
)

logger = logging.getLogger("medace")

# ---------------------------------------------------------------------------
# Model aliases and backend-compatible names
# ---------------------------------------------------------------------------

MODEL_ALIASES: dict[str, str] = {
    "oss-120b": "openai/gpt-oss-120b",
    "oss-20b": "openai/gpt-oss-20b",
    "llama3.1-8b": "meta-llama/Llama-3.1-8B-Instruct",
}
SUPPORTED_ALIASES = tuple(MODEL_ALIASES.keys())


FORM_ORDER = [
    "demographics_birth_history_and_admission",
    "nicu_course",
    "nicu_diagnoses",
    "genetic_tests",
    "patient_phenotypes",
    "genetic_diagnoses",
    "followup_patient_services",
    "vital_status",
    "instrument_for_calculations",
]

GENETICS_BASE_CONTEXT_COLUMNS = [
    "mrn_id",
    "redcap_repeat_instrument",
    "redcap_repeat_instance",
    "study_id",
    "dob",
    "name",
    "sex",
    "nicugt",
    "testperf_nicupost",
    "nicudx",
    "nicudxtype",
]

GENETIC_TEST_ROW_COLUMNS = [
    "type",
    "type_other",
    "reanalysis",
    "testname",
    "geneticrec",
    "resultavailable",
    "file",
    "decline",
    "lab",
    "lab_other",
    "order",
    "returndate",
    "tat",
    "tat_180___1",
    "timeframe",
    "analysis",
    "consent_second",
    "findings",
    "order_age",
    "return_age",
    "time_admit",
]
for _idx in range(1, 21):
    suffix = "" if _idx == 1 else f"_{_idx}"
    GENETIC_TEST_ROW_COLUMNS.extend(
        [
            f"genelocus{suffix}",
            f"dna{suffix}",
            f"protein{suffix}",
            f"transcript{suffix}",
            f"dose{suffix}",
            f"roh{suffix}",
            f"labclass{suffix}",
            f"zygosity{suffix}",
            f"inheritance{suffix}",
            f"segregation{suffix}",
            f"findingclass{suffix}",
            f"second{suffix}",
            f"findingcomments{suffix}",
        ]
    )
GENETIC_TEST_ROW_COLUMNS.extend(
    [
        "testdx",
        "refseq",
        "testcomments",
        "order_post_death",
        "return_post_death",
        "missing_gt",
        "genetic_tests_complete",
    ]
)

PHENOTYPE_ROW_COLUMNS = [
    "hpo_date",
    "hpo_terms",
    "patient_phenotypes_complete",
]

DIAGNOSIS_ROW_COLUMNS: list[str] = []
for _idx in range(1, 4):
    suffix = "" if _idx == 1 else f"_{_idx}"
    DIAGNOSIS_ROW_COLUMNS.extend(
        [
            f"dxname{suffix}",
            f"dx_test{suffix}",
            f"dx_test_other{suffix}",
            f"dxgenelocus{suffix}",
            f"dxdate{suffix}",
            f"dxfhx{suffix}",
            f"dxomim{suffix}",
            f"dxorpha{suffix}",
            f"dxinfo_other{suffix}",
        ]
    )
    if _idx < 3:
        DIAGNOSIS_ROW_COLUMNS.append(f"dxmore{suffix}")
DIAGNOSIS_ROW_COLUMNS.extend(["diagnoses_note", "genetic_diagnoses_complete"])

GENETICS_ONLY_HEADER = (
    GENETICS_BASE_CONTEXT_COLUMNS
    + GENETIC_TEST_ROW_COLUMNS
    + PHENOTYPE_ROW_COLUMNS
    + DIAGNOSIS_ROW_COLUMNS
)

# ---------------------------------------------------------------------------
# REDCap code maps
# ---------------------------------------------------------------------------

SEX_CODES = {
    "1": Sex.male.value,
    "2": Sex.female.value,
    "95": Sex.other.value,
}

TEST_TYPE_CODES = {
    "1": TestType.fish.value,
    "2": TestType.karyotype.value,
    "3": TestType.microarray.value,
    "4": TestType.targeted_variant.value,
    "5": TestType.single_gene_seq_del_dup.value,
    "6": TestType.single_gene_repeat.value,
    "7": TestType.panel_or_exome_slice.value,
    "8": TestType.exome.value,
    "9": TestType.genome.value,
    "10": TestType.mitochondrial.value,
    "11": TestType.imprinting_methylation.value,
    "95": TestType.other.value,
    "99": TestType.not_reported.value,
}
TEST_TYPE_ALIASES = {
    "Karyotype (Chromosome Analysis)": TestType.karyotype.value,
    "Chromosome Analysis": TestType.karyotype.value,
    "Chromosomal Microarray": TestType.microarray.value,
    "CMA": TestType.microarray.value,
    "Microarray (CMA)": TestType.microarray.value,
    "Whole Exome Sequencing": TestType.exome.value,
    "Whole Genome Sequencing": TestType.genome.value,
    "Rapid Whole Genome Sequencing": TestType.genome.value,
    "Rapid Genome Sequencing": TestType.genome.value,
    "rWGS": TestType.genome.value,
}

TEST_LAB_CODES = {
    "1": TestLab.ambry.value,
    "2": TestLab.baylor.value,
    "3": TestLab.blueprint.value,
    "4": TestLab.egl_eurofins.value,
    "5": TestLab.genedx.value,
    "6": TestLab.integrated_genetics.value,
    "7": TestLab.invitae.value,
    "8": TestLab.iu_cytogenetics.value,
    "9": TestLab.iu_molecular.value,
    "10": TestLab.iu_ngs.value,
    "11": TestLab.labcorp.value,
    "12": TestLab.prevention.value,
    "13": TestLab.quest.value,
    "14": TestLab.iu_diagnostic_genomics.value,
    "95": TestLab.other.value,
    "99": TestLab.not_reported.value,
}

ANALYSIS_CODES = {
    "0": AnalysisType.proband_only.value,
    "1": AnalysisType.proband_with_segregation.value,
    "2": AnalysisType.duo.value,
    "3": AnalysisType.trio.value,
    "95": AnalysisType.other.value,
    "99": AnalysisType.not_reported.value,
}

SECONDARY_CODES = {
    "1": SecondaryFindings.opt_in.value,
    "0": SecondaryFindings.opt_out.value,
    "99": SecondaryFindings.not_reported.value,
}

LABCLASS_CODES = {
    "1": LabClassification.pathogenic.value,
    "2": LabClassification.likely_pathogenic.value,
    "3": LabClassification.vus.value,
    "6": LabClassification.abnormal_nos.value,
    "4": LabClassification.likely_benign.value,
    "5": LabClassification.benign.value,
    "0": LabClassification.normal_nos.value,
    "95": LabClassification.other.value,
    "99": LabClassification.not_reported.value,
}
LABCLASS_ALIASES = {"VUS": LabClassification.vus.value}

DOSAGE_CODES = {
    "0": Dosage.x0.value,
    "1": Dosage.x1.value,
    "2": Dosage.x2.value,
    "3": Dosage.x3.value,
    "4": Dosage.x4.value,
    "5": Dosage.copy_loss_nos.value,
    "6": Dosage.copy_gain_nos.value,
    "7": Dosage.deletion.value,
    "8": Dosage.duplication.value,
    "95": Dosage.other.value,
    "99": Dosage.not_reported.value,
}

ZYGOSITY_CODES = {
    "1": Zygosity.heterozygous.value,
    "2": Zygosity.homozygous.value,
    "3": Zygosity.hemizygous.value,
    "95": Zygosity.other.value,
    "99": Zygosity.not_reported.value,
}

INHERITANCE_CODES = {
    "1": InheritancePattern.autosomal_dominant.value,
    "2": InheritancePattern.autosomal_recessive.value,
    "3": InheritancePattern.x_linked.value,
    "4": InheritancePattern.ad_or_ar.value,
    "95": InheritancePattern.other.value,
    "99": InheritancePattern.not_reported.value,
}

SEGREGATION_CODES = {
    "0": Segregation.de_novo.value,
    "1": Segregation.maternal.value,
    "2": Segregation.paternal.value,
    "3": Segregation.not_maternal.value,
    "4": Segregation.not_paternal.value,
    "5": Segregation.maternal_and_paternal.value,
    "95": Segregation.other.value,
    "99": Segregation.not_reported.value,
}

FINDINGCLASS_CODES = {
    "1": FindingClassification.primary.value,
    "2": FindingClassification.secondary.value,
    "3": FindingClassification.incidental.value,
    "99": FindingClassification.unknown.value,
}

TESTDX_CODES = {
    "0": TestInterpretationResult.nondiagnostic.value,
    "1": TestInterpretationResult.diagnostic.value,
    "99": TestInterpretationResult.not_reported.value,
}

REFSEQ_CODES = {
    "1": ReferenceGenome.grch37_hg19.value,
    "2": ReferenceGenome.grch38_hg38.value,
    "95": ReferenceGenome.other.value,
    "99": ReferenceGenome.not_reported.value,
}

TIMEFRAME_CODES = {
    "1": TestTimeframe.nicu_stay.value,
    "2": TestTimeframe.pre_nicu_osh.value,
    "0": TestTimeframe.prenatal.value,
    "3": TestTimeframe.post_nicu_discharge.value,
    "4": TestTimeframe.postmortem.value,
    "95": TestTimeframe.other_unknown.value,
}

ERADX_CODES = {
    "1": DiagnosisStudyPeriod.pre_nicu_stay.value,
    "2": DiagnosisStudyPeriod.nicu_stay.value,
    "3": DiagnosisStudyPeriod.year1_post_nicu.value,
    "4": DiagnosisStudyPeriod.year2_post_nicu.value,
    "95": DiagnosisStudyPeriod.other.value,
    "99": DiagnosisStudyPeriod.unknown.value,
}

NICUDXTYPE_CODES = {
    "0": NICUDiagnosisType.none_unknown.value,
    "1": NICUDiagnosisType.cytogenetic.value,
    "2": NICUDiagnosisType.molecular.value,
    "3": NICUDiagnosisType.two_or_more.value,
    "99": NICUDiagnosisType.not_reported.value,
}


# ---------------------------------------------------------------------------
# Generic parsing/formatting helpers
# ---------------------------------------------------------------------------


def resolve_model(model_name: str) -> str:
    return MODEL_ALIASES.get(model_name, model_name)


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, Enum):
        text = str(value.value).strip()
    else:
        text = str(value).strip()
    if not text:
        return None
    if text.lower() in {"nan", "none", "null"}:
        return None
    return text


def normalize_token(value: Any) -> str:
    text = clean_text(value)
    if text is None:
        return ""
    return " ".join(text.lower().split())


def parse_int(value: Any) -> Optional[int]:
    text = clean_text(value)
    if text is None:
        return None
    try:
        if "." in text:
            return int(float(text))
        return int(text)
    except ValueError:
        return None


def parse_boolish(value: Any) -> Optional[bool]:
    text = normalize_token(value)
    if not text:
        return None
    if text in {"1", "yes", "y", "true", "t"}:
        return True
    if text in {
        "0",
        "no",
        "n",
        "false",
        "f",
        "no/unknown",
        "unknown",
        "not reported",
        "not reported/unknown",
        "n/a",
        "not applicable",
    }:
        return False
    return None


def encode_boolish(
    value: Optional[bool], *, true_code: str = "1", false_code: str = "0"
) -> str:
    if value is None:
        return ""
    return true_code if value else false_code


def lower_lookup(
    code_map: dict[str, str], aliases: Optional[dict[str, str]] = None
) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for code, canonical in code_map.items():
        lookup[normalize_token(code)] = canonical
        lookup[normalize_token(canonical)] = canonical
    for alias, canonical in (aliases or {}).items():
        lookup[normalize_token(alias)] = canonical
    return lookup


_CACHE_LOOKUPS: dict[str, dict[str, str]] = {}


def decode_choice(
    value: Any,
    code_map: dict[str, str],
    aliases: Optional[dict[str, str]] = None,
) -> Optional[str]:
    text = clean_text(value)
    if text is None:
        return None
    cache_key = repr((tuple(code_map.items()), tuple(sorted((aliases or {}).items()))))
    lookup = _CACHE_LOOKUPS.get(cache_key)
    if lookup is None:
        lookup = lower_lookup(code_map, aliases)
        _CACHE_LOOKUPS[cache_key] = lookup
    return lookup.get(normalize_token(text))


def encode_choice(
    value: Any,
    code_map: dict[str, str],
    aliases: Optional[dict[str, str]] = None,
) -> str:
    text = clean_text(value)
    if text is None:
        return ""
    canonical = decode_choice(text, code_map, aliases)
    if canonical is None:
        return text
    for code, candidate in code_map.items():
        if candidate == canonical:
            return code
    return text


def iso_to_mdy(value: Optional[str]) -> str:
    iso = normalize_date(value)
    if not iso:
        return ""
    dt = __import__("datetime").datetime.strptime(iso, "%Y-%m-%d").date()
    return f"{dt.month}/{dt.day}/{dt.year}"


def maybe_json(obj: Any) -> Any:
    if isinstance(obj, BaseModel):
        return obj.model_dump(mode="json")
    return obj


# ---------------------------------------------------------------------------
# PDF / text loading
# ---------------------------------------------------------------------------


def load_report(path: Path, fmt: Optional[str] = None) -> str:
    is_pdf = (fmt == "pdf") if fmt else (path.suffix.lower() == ".pdf")
    if is_pdf:
        return read_pdf(path)
    return path.read_text(encoding="utf-8")


def read_pdf(path: Path) -> str:
    """Extract text from a text-based PDF.

    OCR is intentionally not included here.
    """

    for module_name in ("pymupdf", "fitz"):
        try:
            module = importlib.import_module(module_name)
            document = module.open(str(path))
            try:
                pages: list[str] = []
                for page_number, page in enumerate(document, start=1):
                    if hasattr(page, "get_text"):
                        page_text = page.get_text("text")
                    else:  # pragma: no cover - defensive fallback
                        page_text = page.get_text()
                    pages.append(
                        f"=== PAGE {page_number} ===\n{(page_text or '').strip()}"
                    )
            finally:
                document.close()
            text = "\n\n".join(pages).strip()
            if text:
                return text
        except ImportError:
            continue

    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(str(path)) as pdf:
            pages = []
            for page_number, page in enumerate(pdf.pages, start=1):
                page_text = page.extract_text() or ""
                pages.append(f"=== PAGE {page_number} ===\n{page_text.strip()}")
        text = "\n\n".join(pages).strip()
        if text:
            return text
    except ImportError:
        pass

    raise RuntimeError(
        "Could not extract text from the PDF. Install PyMuPDF (`pip install pymupdf`) "
        "or pdfplumber (`pip install pdfplumber`)."
    )


# ---------------------------------------------------------------------------
# Prompt construction and inference backends
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """\
You are a clinical NLP system specializing in pediatric genetic test report extraction.
Read the report and extract structured data into JSON.

{format_instructions}

GLOBAL RULES:
- Extract only what is explicitly stated or clearly implied by the report.
- Use null or [] exactly as instructed by the schema.
- Keep diagnoses linked to the report that supports them.
- Do not create diagnoses for carrier status, isolated ROH/AOH, isolated VUS, negative results, or a lab classification term without a disorder name.
- Preserve exact HGVS, copy-number, coordinate, and cytogenetic nomenclature.
- For dates, output YYYY-MM-DD.
- If the report is truly negative/nondiagnostic with no reportable findings, set test_info.findings=false and findings_list=[] or null.
- For exome/genome-style reports, capture the phenotype/HPO block in patient_phenotypes when it is present.
- Use test_info.testcomments for free-text shorthand result notes such as karyotype nomenclature when useful.
- Output valid JSON only. No markdown fences. No commentary.
- For diagnosis fields eradx, dxage, dxtimeto: always use null. These are computed post-extraction from NICU admission context.
- For patient_phenotypes: only extract phenotype terms that describe THIS patient's observed clinical features (from referral indication, clinical history, or HPO section). Do not extract general syndrome feature lists from the lab interpretation.
"""

USER_PROMPT = "GENETIC TEST REPORT:\n\n{report_text}"


class ChatBackend:
    def generate(
        self,
        messages: Sequence[dict[str, str]],
        *,
        max_new_tokens: int,
        temperature: float,
        top_p: float,
    ) -> str:
        raise NotImplementedError


class HFBackend(ChatBackend):
    def __init__(
        self, model_id: str, *, num_gpus: int = 1, dtype: str = "auto"
    ) -> None:
        try:
            import torch  # type: ignore
            from transformers import AutoModelForCausalLM, AutoTokenizer  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "HF backend requires torch and transformers. Install them or use --backend vllm."
            ) from exc

        self.torch = torch
        self.model_id = model_id
        self.dtype = dtype

        torch_dtype = self._resolve_dtype(dtype)
        max_memory = self._build_max_memory(num_gpus)

        logger.info("loading tokenizer: %s", model_id)
        self.tokenizer = AutoTokenizer.from_pretrained(model_id, trust_remote_code=True)

        model_kwargs = {
            "dtype": torch_dtype,
            "device_map": "auto",
            "trust_remote_code": True,
        }
        if max_memory:
            model_kwargs["max_memory"] = max_memory
            logger.info("HF backend constrained to max_memory=%s", max_memory)

        logger.info("loading HF model: %s", model_id)
        self.model = AutoModelForCausalLM.from_pretrained(model_id, **model_kwargs)
        self.model.eval()

        if self.tokenizer.pad_token_id is None:
            self.tokenizer.pad_token_id = self.tokenizer.eos_token_id

        self.input_device = self._infer_input_device()
        logger.info("HF backend ready")

    def _resolve_dtype(self, dtype: str):
        if dtype == "auto":
            if self.torch.cuda.is_available():
                return self.torch.bfloat16
            return self.torch.float32
        return getattr(self.torch, dtype)

    def _build_max_memory(self, num_gpus: int) -> Optional[dict[Any, str]]:
        if not self.torch.cuda.is_available():
            return None
        available = self.torch.cuda.device_count()
        if available <= 0:
            return None
        requested = min(max(num_gpus, 1), available)
        if requested == available:
            return None
        max_memory: dict[Any, str] = {"cpu": "128GiB"}
        for index in range(requested):
            total_bytes = self.torch.cuda.get_device_properties(index).total_memory
            budget_gib = max(1, int((total_bytes / (1024**3)) * 0.92))
            max_memory[index] = f"{budget_gib}GiB"
        return max_memory

    def _infer_input_device(self):
        for parameter in self.model.parameters():
            device = getattr(parameter, "device", None)
            if device is not None and getattr(device, "type", None) != "meta":
                return device
        return self.torch.device("cpu")

    @staticmethod
    def _fallback_chat_template(messages: Sequence[dict[str, str]]) -> str:
        chunks: list[str] = []
        for message in messages:
            role = message["role"].upper()
            chunks.append(f"{role}:\n{message['content']}")
        chunks.append("ASSISTANT:\n")
        return "\n\n".join(chunks)

    def generate(
        self,
        messages: Sequence[dict[str, str]],
        *,
        max_new_tokens: int,
        temperature: float,
        top_p: float,
    ) -> str:
        if hasattr(self.tokenizer, "apply_chat_template"):
            prompt = self.tokenizer.apply_chat_template(
                list(messages),
                tokenize=False,
                add_generation_prompt=True,
            )
        else:
            prompt = self._fallback_chat_template(messages)

        encoded = self.tokenizer(prompt, return_tensors="pt")
        encoded = {key: value.to(self.input_device) for key, value in encoded.items()}

        do_sample = temperature > 0
        generate_kwargs: dict[str, Any] = {
            **encoded,
            "max_new_tokens": max_new_tokens,
            "do_sample": do_sample,
            "pad_token_id": self.tokenizer.pad_token_id,
        }
        if do_sample:
            generate_kwargs["temperature"] = temperature
            generate_kwargs["top_p"] = top_p

        with self.torch.no_grad():
            outputs = self.model.generate(**generate_kwargs)

        generated = outputs[0, encoded["input_ids"].shape[1] :]
        return self.tokenizer.decode(generated, skip_special_tokens=True).strip()


class VLLMBackend(ChatBackend):
    def __init__(
        self,
        model_id: str,
        *,
        base_url: str,
        read_timeout: int = 1800,
        api_key: Optional[str] = None,
    ) -> None:
        self.model_id = model_id
        self.base_url = normalize_base_url(base_url)
        self.read_timeout = read_timeout
        self.api_key = api_key
        self.models_url = f"{self.base_url}/models"
        self.chat_url = f"{self.base_url}/chat/completions"
        self._probe_server()

    def _headers(self) -> dict[str, str]:
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        return headers

    def _request_json(
        self,
        method: str,
        url: str,
        payload: Optional[dict[str, Any]] = None,
    ) -> dict[str, Any]:
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            url, data=data, headers=self._headers(), method=method
        )
        try:
            with urllib.request.urlopen(request, timeout=self.read_timeout) as response:
                body = response.read().decode("utf-8")
                return json.loads(body)
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"HTTP {exc.code} from {url}: {body}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"Could not reach vLLM server at {url}: {exc}") from exc

    def _probe_server(self) -> None:
        logger.info("probing vLLM server: %s", self.models_url)
        payload = self._request_json("GET", self.models_url)
        available_models = [
            item.get("id") for item in payload.get("data", []) if isinstance(item, dict)
        ]
        if available_models:
            logger.info("vLLM server ready. Models: %s", ", ".join(available_models))
            if self.model_id not in available_models:
                logger.warning(
                    "requested model '%s' not listed by the server; continuing anyway",
                    self.model_id,
                )

    @staticmethod
    def _coerce_message_content(content: Any) -> str:
        if isinstance(content, str):
            return content
        if isinstance(content, list):
            parts: list[str] = []
            for item in content:
                if isinstance(item, dict) and item.get("type") == "text":
                    parts.append(str(item.get("text", "")))
                else:
                    parts.append(str(item))
            return "\n".join(part for part in parts if part)
        return str(content)

    def generate(
        self,
        messages: Sequence[dict[str, str]],
        *,
        max_new_tokens: int,
        temperature: float,
        top_p: float,
    ) -> str:
        payload: dict[str, Any] = {
            "model": self.model_id,
            "messages": list(messages),
            "max_tokens": max_new_tokens,
            "temperature": max(temperature, 0.0),
        }
        if temperature > 0:
            payload["top_p"] = top_p
        response = self._request_json("POST", self.chat_url, payload)
        try:
            message = response["choices"][0]["message"]["content"]
        except (KeyError, IndexError, TypeError) as exc:  # pragma: no cover - defensive
            raise RuntimeError(f"Unexpected vLLM response payload: {response}") from exc
        return self._coerce_message_content(message).strip()


def build_messages(
    report_text: str, model_cls: Type[BaseModel]
) -> list[dict[str, str]]:
    format_instructions = build_format_instructions(model_cls)
    return [
        {
            "role": "system",
            "content": SYSTEM_PROMPT.format(format_instructions=format_instructions),
        },
        {"role": "user", "content": USER_PROMPT.format(report_text=report_text)},
    ]


# ---------------------------------------------------------------------------
# Extraction loop
# ---------------------------------------------------------------------------


def strip_json_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```json"):
        text = text[7:]
    elif text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    return text.strip()


def maybe_extract_json_object(text: str) -> str:
    text = strip_json_fences(text)
    if text.startswith("{") and text.endswith("}"):
        return text
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start : end + 1]
    return text


def extract(
    report_text: str,
    backend: ChatBackend,
    *,
    model_cls: Optional[Type[BaseModel]] = None,
    temperature: float = 0.0,
    top_p: float = 1.0,
    max_new_tokens: int = 32768,
    max_retries: int = 3,
) -> BaseModel:
    if model_cls is None:
        model_cls = get_extraction_model()

    base_messages = build_messages(report_text, model_cls)
    current_messages = list(base_messages)
    last_error: Optional[Exception] = None

    for attempt in range(1, max_retries + 1):
        logger.info("extraction attempt %d/%d", attempt, max_retries)
        raw = backend.generate(
            current_messages,
            max_new_tokens=max_new_tokens,
            temperature=temperature,
            top_p=top_p,
        )
        logger.debug("raw model response:\n%s", raw)

        try:
            payload = json.loads(maybe_extract_json_object(raw))
        except json.JSONDecodeError as exc:
            last_error = exc
            current_messages = list(base_messages) + [
                {"role": "assistant", "content": raw},
                {
                    "role": "user",
                    "content": (
                        "Your previous response was not valid JSON. "
                        f"JSON parsing error: {exc}. Return corrected JSON only."
                    ),
                },
            ]
            continue

        try:
            return model_cls.model_validate(payload)
        except ValidationError as exc:
            last_error = exc
            logger.warning("schema validation failed (%d errors)", exc.error_count())
            logger.debug("validation errors: %s", exc.errors())
            current_messages = list(base_messages) + [
                {"role": "assistant", "content": raw},
                {
                    "role": "user",
                    "content": (
                        "Your previous JSON did not validate against the schema. "
                        f"Validation error details:\n{exc}\n\n"
                        "Fix the JSON and return corrected JSON only."
                    ),
                },
            ]
            continue

    assert last_error is not None
    raise last_error


# ---------------------------------------------------------------------------
# REDCap data dictionary helpers
# ---------------------------------------------------------------------------


def parse_redcap_choices(choices: Optional[str]) -> list[tuple[str, str]]:
    text = clean_text(choices)
    if text is None:
        return []
    parts = [part.strip() for part in text.split("|")]
    parsed: list[tuple[str, str]] = []
    for part in parts:
        if not part:
            continue
        if "," in part:
            code, label = part.split(",", 1)
            parsed.append((code.strip(), label.strip()))
        else:
            parsed.append((part, part))
    return parsed


def build_redcap_column_order_from_dictionary(data_dictionary_path: Path) -> list[str]:
    import openpyxl

    workbook = openpyxl.load_workbook(
        data_dictionary_path, read_only=True, data_only=True
    )
    worksheet = workbook[workbook.sheetnames[0]]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {name: idx for idx, name in enumerate(header)}

    def get(row: Sequence[Any], name: str) -> Any:
        return row[header_index[name]] if name in header_index else None

    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    nonblank_rows = [row for row in rows if get(row, "Variable / Field Name")]
    if not nonblank_rows:
        return GENETICS_ONLY_HEADER.copy()

    record_id = clean_text(get(nonblank_rows[0], "Variable / Field Name")) or "mrn_id"
    column_order: list[str] = [
        record_id,
        "redcap_repeat_instrument",
        "redcap_repeat_instance",
    ]

    current_form: Optional[str] = None
    form_columns: list[str] = []

    def flush_form(form_name: Optional[str]) -> None:
        if not form_name:
            return
        column_order.extend(form_columns)
        status_column = f"{form_name}_complete"
        if status_column not in column_order:
            column_order.append(status_column)

    for row in nonblank_rows:
        variable = clean_text(get(row, "Variable / Field Name"))
        form_name = clean_text(get(row, "Form Name"))
        field_type = clean_text(get(row, "Field Type"))
        choices = clean_text(get(row, "Choices, Calculations, OR Slider Labels"))
        if variable is None or form_name is None:
            continue
        if current_form is None:
            current_form = form_name
        if form_name != current_form:
            flush_form(current_form)
            form_columns = []
            current_form = form_name

        if variable == record_id or field_type == "descriptive":
            continue
        if field_type == "checkbox":
            for code, _label in parse_redcap_choices(choices):
                form_columns.append(f"{variable}___{code}")
        else:
            form_columns.append(variable)

    flush_form(current_form)
    workbook.close()
    return column_order


# ---------------------------------------------------------------------------
# Tabular IO helpers
# ---------------------------------------------------------------------------


def detect_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return dialect.delimiter
    except csv.Error:
        if "\t" in sample and sample.count("\t") >= sample.count(","):
            return "\t"
        return ","


def read_rows_from_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        workbook.close()
        return [], []
    header = [str(cell).strip() if cell is not None else "" for cell in header_row]
    rows: list[dict[str, str]] = []
    for raw_row in iterator:
        row_dict: dict[str, str] = {}
        for index, column in enumerate(header):
            value = raw_row[index] if index < len(raw_row) else None
            row_dict[column] = "" if value is None else str(value)
        if any(clean_text(value) is not None for value in row_dict.values()):
            rows.append(row_dict)
    workbook.close()
    return header, rows


def read_rows_from_delimited(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    sample = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = detect_delimiter(sample[:4096])
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        header = list(reader.fieldnames or [])
        rows = []
        for row in reader:
            if row is None:
                continue
            cleaned = {
                key: ("" if value is None else value) for key, value in row.items()
            }
            if any(clean_text(value) is not None for value in cleaned.values()):
                rows.append(cleaned)
    return header, rows


def read_tabular_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_rows_from_xlsx(path)
    return read_rows_from_delimited(path)


def write_rows_to_delimited(
    path: Path,
    header: list[str],
    rows: list[dict[str, Any]],
) -> None:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=header, delimiter=delimiter, extrasaction="ignore"
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in header})


# ---------------------------------------------------------------------------
# REDCap import helpers
# ---------------------------------------------------------------------------


def blank_or_none(value: Any) -> str:
    text = clean_text(value)
    return text or ""


def copy_row_for_json(row: dict[str, Any]) -> dict[str, Any]:
    return {
        str(key): ("" if value is None else str(value)) for key, value in row.items()
    }


def first_nonblank(*values: Any) -> Optional[str]:
    for value in values:
        text = clean_text(value)
        if text is not None:
            return text
    return None


def extract_patient_from_any_row(row: dict[str, Any]) -> PatientDemographics:
    return PatientDemographics(
        mrn_id=clean_text(row.get("mrn_id")),
        study_id=clean_text(row.get("study_id")),
        name=clean_text(row.get("name")),
        dob=clean_text(row.get("dob")),
        sex=decode_choice(row.get("sex"), SEX_CODES),
    )


def extract_nicu_flags(base_row: Optional[dict[str, Any]]) -> Optional[dict[str, Any]]:
    if not base_row:
        return None
    flags = {
        "nicugt": parse_boolish(base_row.get("nicugt")),
        "testperf_nicupost": parse_boolish(base_row.get("testperf_nicupost")),
        "nicudx": parse_boolish(base_row.get("nicudx")),
        "nicudxtype": decode_choice(base_row.get("nicudxtype"), NICUDXTYPE_CODES),
        "counsel": parse_boolish(base_row.get("counsel")),
        "counsel_date": (
            normalize_date(base_row.get("counsel_date"))
            if clean_text(base_row.get("counsel_date"))
            else None
        ),
        "nicueval": parse_boolish(base_row.get("nicueval")),
        "nicueval_date": (
            normalize_date(base_row.get("nicueval_date"))
            if clean_text(base_row.get("nicueval_date"))
            else None
        ),
        "nicueval_age": parse_int(base_row.get("nicueval_age")),
        "nicueval_timeto": parse_int(base_row.get("nicueval_timeto")),
        "neval_time": clean_text(base_row.get("neval_time")),
    }
    return {key: value for key, value in flags.items() if value is not None}


def parse_finding_from_row(row: dict[str, Any], index: int) -> Optional[GeneticFinding]:
    suffix = "" if index == 1 else f"_{index}"
    keys = [
        f"genelocus{suffix}",
        f"dna{suffix}",
        f"protein{suffix}",
        f"transcript{suffix}",
        f"dose{suffix}",
        f"roh{suffix}",
        f"labclass{suffix}",
        f"zygosity{suffix}",
        f"inheritance{suffix}",
        f"segregation{suffix}",
        f"findingclass{suffix}",
        f"findingcomments{suffix}",
    ]
    if not any(clean_text(row.get(key)) is not None for key in keys):
        return None
    return GeneticFinding(
        genelocus=clean_text(row.get(f"genelocus{suffix}")),
        dna=clean_text(row.get(f"dna{suffix}")),
        protein=clean_text(row.get(f"protein{suffix}")),
        transcript=clean_text(row.get(f"transcript{suffix}")),
        dose=decode_choice(row.get(f"dose{suffix}"), DOSAGE_CODES),
        roh=parse_boolish(row.get(f"roh{suffix}")),
        labclass=decode_choice(
            row.get(f"labclass{suffix}"), LABCLASS_CODES, LABCLASS_ALIASES
        ),
        zygosity=decode_choice(row.get(f"zygosity{suffix}"), ZYGOSITY_CODES),
        inheritance=decode_choice(row.get(f"inheritance{suffix}"), INHERITANCE_CODES),
        segregation=decode_choice(row.get(f"segregation{suffix}"), SEGREGATION_CODES),
        findingclass=decode_choice(
            row.get(f"findingclass{suffix}"), FINDINGCLASS_CODES
        ),
        findingcomments=clean_text(row.get(f"findingcomments{suffix}")),
    )


def parse_genetic_test_row(
    row: dict[str, Any],
    *,
    patient: PatientDemographics,
    repeat_instance: Optional[int],
) -> LinkedGeneticReport:
    findings_list = [
        finding
        for finding in (parse_finding_from_row(row, index) for index in range(1, 21))
        if finding is not None
    ]

    report = LinkedGeneticReport(
        patient=patient,
        report_metadata=ReportMetadata(),
        test_info=TestInformation(
            type=decode_choice(row.get("type"), TEST_TYPE_CODES, TEST_TYPE_ALIASES),
            type_other=clean_text(row.get("type_other")),
            reanalysis=parse_boolish(row.get("reanalysis")),
            testname=clean_text(row.get("testname")),
            geneticrec=parse_boolish(row.get("geneticrec")),
            resultavailable=parse_boolish(row.get("resultavailable")),
            file=clean_text(row.get("file")),
            decline=parse_boolish(row.get("decline")),
            lab=decode_choice(row.get("lab"), TEST_LAB_CODES),
            lab_other=clean_text(row.get("lab_other")),
            order=clean_text(row.get("order")),
            returndate=clean_text(row.get("returndate")),
            tat=parse_int(row.get("tat")),
            timeframe=decode_choice(row.get("timeframe"), TIMEFRAME_CODES),
            analysis=decode_choice(row.get("analysis"), ANALYSIS_CODES),
            consent_second=decode_choice(row.get("consent_second"), SECONDARY_CODES),
            findings=parse_boolish(row.get("findings")),
            order_age=parse_int(row.get("order_age")),
            return_age=parse_int(row.get("return_age")),
            time_admit=parse_int(row.get("time_admit")),
            testcomments=clean_text(row.get("testcomments")),
            order_post_death=clean_text(row.get("order_post_death")),
            return_post_death=clean_text(row.get("return_post_death")),
            missing_gt=clean_text(row.get("missing_gt")),
        ),
        findings_list=findings_list or None,
        interpretation=TestInterpretation(
            testdx=decode_choice(row.get("testdx"), TESTDX_CODES),
            refseq=decode_choice(row.get("refseq"), REFSEQ_CODES),
            interpretation_text=None,
        ),
        patient_phenotypes=None,
        diagnoses=None,
        diagnoses_note=None,
        extraction_confidence=None,
        extraction_notes=None,
        source_repeat_instance=repeat_instance,
        source_row=copy_row_for_json(row),
    )
    return report


def parse_phenotype_row(
    row: dict[str, Any], *, repeat_instance: Optional[int]
) -> Optional[LegacyPhenotypeNote]:
    if (
        clean_text(row.get("hpo_date")) is None
        and clean_text(row.get("hpo_terms")) is None
    ):
        return None
    note = LegacyPhenotypeNote(
        hpo_date=clean_text(row.get("hpo_date")),
        hpo_terms=clean_text(row.get("hpo_terms")),
        parsed_terms=None,
        source_repeat_instance=repeat_instance,
        source_row=copy_row_for_json(row),
    )
    return note


def parse_diagnosis_from_slot(
    row: dict[str, Any], index: int
) -> Optional[GeneticDiagnosis]:
    suffix = "" if index == 1 else f"_{index}"
    keys = [
        f"dxname{suffix}",
        f"dx_test{suffix}",
        f"dx_test_other{suffix}",
        f"dxgenelocus{suffix}",
        f"dxdate{suffix}",
        f"dxfhx{suffix}",
        f"dxomim{suffix}",
        f"dxorpha{suffix}",
        f"dxinfo_other{suffix}",
    ]
    if not any(clean_text(row.get(key)) is not None for key in keys):
        return None
    return GeneticDiagnosis(
        dxname=clean_text(row.get(f"dxname{suffix}")),
        dx_test=decode_choice(
            row.get(f"dx_test{suffix}"), TEST_TYPE_CODES, TEST_TYPE_ALIASES
        ),
        dx_test_other=clean_text(row.get(f"dx_test_other{suffix}")),
        dxgenelocus=clean_text(row.get(f"dxgenelocus{suffix}")),
        dxdate=clean_text(row.get(f"dxdate{suffix}")),
        dxfhx=parse_boolish(row.get(f"dxfhx{suffix}")),
        dxomim=clean_text(row.get(f"dxomim{suffix}")),
        dxorpha=clean_text(row.get(f"dxorpha{suffix}")),
        dxinfo_other=clean_text(row.get(f"dxinfo_other{suffix}")),
    )


def parse_diagnosis_group_row(
    row: dict[str, Any], *, repeat_instance: Optional[int]
) -> Optional[GeneticDiagnosisGroup]:
    diagnoses = [
        diagnosis
        for diagnosis in (
            parse_diagnosis_from_slot(row, index) for index in range(1, 4)
        )
        if diagnosis is not None
    ]
    note = clean_text(row.get("diagnoses_note"))
    if not diagnoses and note is None:
        return None
    return GeneticDiagnosisGroup(
        diagnoses=diagnoses,
        diagnoses_note=note,
        source_repeat_instance=repeat_instance,
        source_row=copy_row_for_json(row),
    )


def report_gene_tokens(report: LinkedGeneticReport) -> set[str]:
    tokens: set[str] = set()
    for finding in report.findings_list or []:
        locus = normalize_token(finding.genelocus)
        if locus:
            tokens.add(locus)
    free_text_bits = [
        report.test_info.testname,
        report.test_info.testcomments,
        report.interpretation.interpretation_text,
        report.diagnoses_note,
    ]
    for bit in free_text_bits:
        norm = normalize_token(bit)
        if norm:
            tokens.add(norm)
    return tokens


def score_diagnosis_group_against_report(
    group: GeneticDiagnosisGroup,
    report: LinkedGeneticReport,
) -> int:
    score = 0
    group_types = {
        diagnosis.dx_test for diagnosis in group.diagnoses if diagnosis.dx_test
    }
    group_dates = {
        diagnosis.dxdate for diagnosis in group.diagnoses if diagnosis.dxdate
    }
    group_loci = {
        normalize_token(diagnosis.dxgenelocus)
        for diagnosis in group.diagnoses
        if diagnosis.dxgenelocus
    }
    report_tokens = report_gene_tokens(report)

    if report.test_info.type and report.test_info.type in group_types:
        score += 3
    if report.test_info.returndate and report.test_info.returndate in group_dates:
        score += 4
    elif report.test_info.order and report.test_info.order in group_dates:
        score += 2
    if group_loci:
        if group_loci & report_tokens:
            score += 2
        elif any(
            locus and any(locus in token for token in report_tokens)
            for locus in group_loci
        ):
            score += 1
    return score


def attach_legacy_rows_to_reports(
    reports: list[LinkedGeneticReport],
    phenotype_rows: list[LegacyPhenotypeNote],
    diagnosis_groups: list[GeneticDiagnosisGroup],
) -> tuple[list[LegacyPhenotypeNote], list[GeneticDiagnosisGroup]]:
    if not reports:
        return phenotype_rows, diagnosis_groups

    if len(reports) == 1:
        report = reports[0]
        report.patient_phenotypes = [
            PatientPhenotypeNote(
                hpo_date=note.hpo_date,
                hpo_terms=note.hpo_terms,
                parsed_terms=note.parsed_terms,
            )
            for note in phenotype_rows
        ] or None
        report.linked_patient_phenotypes = phenotype_rows
        linked_diagnoses: list[GeneticDiagnosis] = []
        note_texts: list[str] = []
        for group in diagnosis_groups:
            report.linked_diagnosis_groups.append(group)
            linked_diagnoses.extend(group.diagnoses)
            if group.diagnoses_note:
                note_texts.append(group.diagnoses_note)
        report.diagnoses = linked_diagnoses or None
        report.diagnoses_note = (
            "\n\n".join(note_texts) if note_texts else report.diagnoses_note
        )
        report.link_method = "single_report"
        return [], []

    remaining_phenotypes: list[LegacyPhenotypeNote] = []
    for note in phenotype_rows:
        hpo_date = note.hpo_date
        matches = [
            report
            for report in reports
            if hpo_date
            and hpo_date in {report.test_info.order, report.test_info.returndate}
        ]
        if len(matches) == 1:
            report = matches[0]
            report.linked_patient_phenotypes.append(note)
            current = list(report.patient_phenotypes or [])
            current.append(
                PatientPhenotypeNote(
                    hpo_date=note.hpo_date,
                    hpo_terms=note.hpo_terms,
                    parsed_terms=note.parsed_terms,
                )
            )
            report.patient_phenotypes = current
            report.link_method = report.link_method or "date_match"
        else:
            remaining_phenotypes.append(note)

    remaining_diagnosis_groups: list[GeneticDiagnosisGroup] = []
    for group in diagnosis_groups:
        scores = [
            (score_diagnosis_group_against_report(group, report), report)
            for report in reports
        ]
        scores.sort(key=lambda item: item[0], reverse=True)
        if scores and scores[0][0] >= 3:
            top_score = scores[0][0]
            tied = [report for score, report in scores if score == top_score]
            if len(tied) == 1:
                report = tied[0]
                report.linked_diagnosis_groups.append(group)
                current_dx = list(report.diagnoses or [])
                current_dx.extend(group.diagnoses)
                report.diagnoses = current_dx or None
                note_bits = [
                    bit for bit in [report.diagnoses_note, group.diagnoses_note] if bit
                ]
                report.diagnoses_note = "\n\n".join(note_bits) if note_bits else None
                report.link_method = report.link_method or "heuristic_score"
                continue
        remaining_diagnosis_groups.append(group)

    return remaining_phenotypes, remaining_diagnosis_groups


def import_redcap_rows_to_cases(
    header: list[str],
    rows: list[dict[str, str]],
    *,
    source_file: Optional[str] = None,
) -> list[LinkedGeneticsCase]:
    grouped: OrderedDict[str, list[tuple[int, dict[str, str]]]] = OrderedDict()
    for index, row in enumerate(rows):
        mrn_id = clean_text(row.get("mrn_id")) or f"__missing_mrn__{index}"
        grouped.setdefault(mrn_id, []).append((index, row))

    cases: list[LinkedGeneticsCase] = []
    for mrn_id, group_rows in grouped.items():
        base_row: Optional[dict[str, Any]] = None
        reports: list[LinkedGeneticReport] = []
        phenotype_rows: list[LegacyPhenotypeNote] = []
        diagnosis_groups: list[GeneticDiagnosisGroup] = []
        other_repeat_rows: list[LegacyRepeatRow] = []
        patient_seed: Optional[PatientDemographics] = None

        for source_index, row in group_rows:
            instrument = clean_text(row.get("redcap_repeat_instrument"))
            repeat_instance = parse_int(row.get("redcap_repeat_instance"))
            if instrument is None:
                base_row = copy_row_for_json(row)
                patient_seed = extract_patient_from_any_row(row)
                continue

            if patient_seed is None:
                patient_seed = extract_patient_from_any_row(row)

            if instrument == "genetic_tests":
                reports.append(
                    parse_genetic_test_row(
                        row,
                        patient=patient_seed,
                        repeat_instance=repeat_instance,
                    )
                )
            elif instrument == "patient_phenotypes":
                note = parse_phenotype_row(row, repeat_instance=repeat_instance)
                if note is not None:
                    phenotype_rows.append(note)
            elif instrument == "genetic_diagnoses":
                group = parse_diagnosis_group_row(row, repeat_instance=repeat_instance)
                if group is not None:
                    diagnosis_groups.append(group)
            else:
                other_repeat_rows.append(
                    LegacyRepeatRow(
                        redcap_repeat_instrument=instrument,
                        redcap_repeat_instance=repeat_instance,
                        row_data=copy_row_for_json(row),
                        source_index=source_index,
                    )
                )

        patient = patient_seed or PatientDemographics(mrn_id=mrn_id)
        if base_row:
            patient_from_base = extract_patient_from_any_row(base_row)
            patient = PatientDemographics(
                mrn_id=first_nonblank(patient_from_base.mrn_id, patient.mrn_id),
                study_id=first_nonblank(patient_from_base.study_id, patient.study_id),
                name=first_nonblank(patient_from_base.name, patient.name),
                patient_name_last=first_nonblank(
                    patient_from_base.patient_name_last, patient.patient_name_last
                ),
                patient_name_first=first_nonblank(
                    patient_from_base.patient_name_first, patient.patient_name_first
                ),
                dob=first_nonblank(patient_from_base.dob, patient.dob),
                sex=patient_from_base.sex or patient.sex,
            )

        remaining_phenotypes, remaining_diagnosis_groups = (
            attach_legacy_rows_to_reports(
                reports,
                phenotype_rows,
                diagnosis_groups,
            )
        )

        cases.append(
            LinkedGeneticsCase(
                patient=patient,
                nicu_flags=extract_nicu_flags(base_row),
                reports=reports,
                unlinked_patient_phenotypes=remaining_phenotypes,
                unlinked_diagnosis_groups=remaining_diagnosis_groups,
                legacy_base_row=base_row,
                legacy_other_repeat_rows=sorted(
                    other_repeat_rows,
                    key=lambda row_obj: (
                        (
                            FORM_ORDER.index(row_obj.redcap_repeat_instrument)
                            if row_obj.redcap_repeat_instrument in FORM_ORDER
                            else len(FORM_ORDER)
                        ),
                        row_obj.redcap_repeat_instance or 0,
                        row_obj.source_index or 0,
                    ),
                ),
                redcap_column_order=header,
                source_file=source_file,
            )
        )

    return cases


# ---------------------------------------------------------------------------
# REDCap export helpers
# ---------------------------------------------------------------------------


def empty_row(header: list[str]) -> dict[str, str]:
    return {column: "" for column in header}


def clear_columns(row: dict[str, Any], columns: Iterable[str]) -> None:
    for column in columns:
        row[column] = ""


def set_if_present(row: dict[str, Any], column: str, value: Any) -> None:
    if column not in row:
        return
    row[column] = "" if value is None else str(value)


def populate_patient_base_fields(
    row: dict[str, Any], patient: PatientDemographics
) -> None:
    set_if_present(row, "mrn_id", patient.mrn_id)
    set_if_present(row, "study_id", patient.study_id)
    set_if_present(row, "dob", iso_to_mdy(patient.dob))
    set_if_present(
        row,
        "name",
        compose_name(
            patient.name, patient.patient_name_last, patient.patient_name_first
        ),
    )
    set_if_present(row, "sex", encode_choice(patient.sex, SEX_CODES))


def choose_instance(preferred: Optional[int], used: set[int]) -> int:
    if preferred is not None and preferred > 0 and preferred not in used:
        used.add(preferred)
        return preferred
    candidate = 1
    while candidate in used:
        candidate += 1
    used.add(candidate)
    return candidate


def preferred_file_name(path: Optional[str]) -> Optional[str]:
    if not path:
        return None
    return Path(path).name


def apply_test_report_to_row(
    row: dict[str, Any],
    report: LinkedGeneticReport,
) -> None:
    clear_columns(row, GENETIC_TEST_ROW_COLUMNS)
    set_if_present(
        row,
        "type",
        encode_choice(report.test_info.type, TEST_TYPE_CODES, TEST_TYPE_ALIASES),
    )
    set_if_present(row, "type_other", report.test_info.type_other)
    set_if_present(row, "reanalysis", encode_boolish(report.test_info.reanalysis))
    set_if_present(row, "testname", report.test_info.testname)
    set_if_present(row, "geneticrec", encode_boolish(report.test_info.geneticrec))
    resultavailable = report.test_info.resultavailable
    if resultavailable is None:
        resultavailable = (
            True if report.test_info.returndate or report.test_info.file else None
        )
    set_if_present(row, "resultavailable", encode_boolish(resultavailable))
    set_if_present(row, "file", preferred_file_name(report.test_info.file))
    set_if_present(row, "decline", encode_boolish(report.test_info.decline))
    set_if_present(row, "lab", encode_choice(report.test_info.lab, TEST_LAB_CODES))
    set_if_present(row, "lab_other", report.test_info.lab_other)
    set_if_present(row, "order", iso_to_mdy(report.test_info.order))
    set_if_present(row, "returndate", iso_to_mdy(report.test_info.returndate))
    set_if_present(row, "tat", report.test_info.tat)
    if report.test_info.tat is not None:
        set_if_present(row, "tat_180___1", "1" if report.test_info.tat <= 180 else "0")
    set_if_present(
        row, "timeframe", encode_choice(report.test_info.timeframe, TIMEFRAME_CODES)
    )
    set_if_present(
        row, "analysis", encode_choice(report.test_info.analysis, ANALYSIS_CODES)
    )
    set_if_present(
        row,
        "consent_second",
        encode_choice(report.test_info.consent_second, SECONDARY_CODES),
    )
    set_if_present(row, "findings", encode_boolish(report.test_info.findings))
    set_if_present(row, "order_age", report.test_info.order_age)
    set_if_present(row, "return_age", report.test_info.return_age)
    set_if_present(row, "time_admit", report.test_info.time_admit)

    findings = list(report.findings_list or [])
    if len(findings) > 20:
        raise ValueError(
            f"Report '{report.test_info.testname or report.source_repeat_instance}' has {len(findings)} findings; REDCap genetic_tests supports at most 20 per row."
        )

    for index in range(1, 21):
        suffix = "" if index == 1 else f"_{index}"
        finding = findings[index - 1] if index <= len(findings) else None
        set_if_present(
            row, f"genelocus{suffix}", finding.genelocus if finding else None
        )
        set_if_present(row, f"dna{suffix}", finding.dna if finding else None)
        set_if_present(row, f"protein{suffix}", finding.protein if finding else None)
        set_if_present(
            row, f"transcript{suffix}", finding.transcript if finding else None
        )
        set_if_present(
            row,
            f"dose{suffix}",
            encode_choice(finding.dose if finding else None, DOSAGE_CODES),
        )
        set_if_present(
            row, f"roh{suffix}", encode_boolish(finding.roh if finding else None)
        )
        set_if_present(
            row,
            f"labclass{suffix}",
            encode_choice(
                finding.labclass if finding else None, LABCLASS_CODES, LABCLASS_ALIASES
            ),
        )
        set_if_present(
            row,
            f"zygosity{suffix}",
            encode_choice(finding.zygosity if finding else None, ZYGOSITY_CODES),
        )
        set_if_present(
            row,
            f"inheritance{suffix}",
            encode_choice(finding.inheritance if finding else None, INHERITANCE_CODES),
        )
        set_if_present(
            row,
            f"segregation{suffix}",
            encode_choice(finding.segregation if finding else None, SEGREGATION_CODES),
        )
        set_if_present(
            row,
            f"findingclass{suffix}",
            encode_choice(
                finding.findingclass if finding else None, FINDINGCLASS_CODES
            ),
        )
        has_next = index < len(findings)
        set_if_present(row, f"second{suffix}", "1" if has_next else "0")
        set_if_present(
            row,
            f"findingcomments{suffix}",
            finding.findingcomments if finding else None,
        )

    set_if_present(
        row, "testdx", encode_choice(report.interpretation.testdx, TESTDX_CODES)
    )
    set_if_present(
        row, "refseq", encode_choice(report.interpretation.refseq, REFSEQ_CODES)
    )
    testcomments = (
        report.test_info.testcomments or report.interpretation.interpretation_text
    )
    set_if_present(row, "testcomments", testcomments)
    set_if_present(row, "order_post_death", report.test_info.order_post_death)
    set_if_present(row, "return_post_death", report.test_info.return_post_death)
    set_if_present(row, "missing_gt", report.test_info.missing_gt)
    set_if_present(row, "genetic_tests_complete", "2")


def apply_phenotype_to_row(
    row: dict[str, Any],
    note: PatientPhenotypeNote,
    fallback_order_date: Optional[str] = None,
) -> None:
    clear_columns(row, PHENOTYPE_ROW_COLUMNS)
    hpo_date = note.hpo_date or fallback_order_date
    set_if_present(row, "hpo_date", iso_to_mdy(hpo_date))
    set_if_present(row, "hpo_terms", note.hpo_terms)
    set_if_present(row, "patient_phenotypes_complete", "2")


def diagnosis_with_report_defaults(
    diagnosis: GeneticDiagnosis,
    report: Optional[LinkedGeneticReport] = None,
) -> GeneticDiagnosis:
    updates: dict[str, Any] = {}
    if report is not None:
        if diagnosis.dx_test is None and report.test_info.type is not None:
            updates["dx_test"] = report.test_info.type
        if diagnosis.dxdate is None and report.test_info.returndate is not None:
            updates["dxdate"] = report.test_info.returndate
    if not updates:
        return diagnosis
    return diagnosis.model_copy(update=updates)


def chunk_diagnoses(
    diagnoses: list[GeneticDiagnosis], size: int = 3
) -> list[list[GeneticDiagnosis]]:
    return [diagnoses[index : index + size] for index in range(0, len(diagnoses), size)]


def apply_diagnosis_group_to_row(
    row: dict[str, Any],
    diagnoses: list[GeneticDiagnosis],
    diagnoses_note: Optional[str],
) -> None:
    clear_columns(row, DIAGNOSIS_ROW_COLUMNS)
    for index in range(1, 4):
        suffix = "" if index == 1 else f"_{index}"
        diagnosis = diagnoses[index - 1] if index <= len(diagnoses) else None
        set_if_present(row, f"dxname{suffix}", diagnosis.dxname if diagnosis else None)
        set_if_present(
            row,
            f"dx_test{suffix}",
            encode_choice(
                diagnosis.dx_test if diagnosis else None,
                TEST_TYPE_CODES,
                TEST_TYPE_ALIASES,
            ),
        )
        set_if_present(
            row,
            f"dx_test_other{suffix}",
            diagnosis.dx_test_other if diagnosis else None,
        )
        set_if_present(
            row, f"dxgenelocus{suffix}", diagnosis.dxgenelocus if diagnosis else None
        )
        set_if_present(
            row, f"dxdate{suffix}", iso_to_mdy(diagnosis.dxdate if diagnosis else None)
        )
        set_if_present(
            row,
            f"dxfhx{suffix}",
            encode_boolish(diagnosis.dxfhx if diagnosis else None),
        )
        set_if_present(row, f"dxomim{suffix}", diagnosis.dxomim if diagnosis else None)
        set_if_present(
            row, f"dxorpha{suffix}", diagnosis.dxorpha if diagnosis else None
        )
        set_if_present(
            row, f"dxinfo_other{suffix}", diagnosis.dxinfo_other if diagnosis else None
        )
        if index < 3:
            has_next = index < len(diagnoses)
            set_if_present(row, f"dxmore{suffix}", "1" if has_next else "0")
    set_if_present(row, "diagnoses_note", diagnoses_note)
    set_if_present(row, "genetic_diagnoses_complete", "2")


def synthesize_linked_phenotype_rows(
    report: LinkedGeneticReport,
) -> list[LegacyPhenotypeNote]:
    notes = list(report.linked_patient_phenotypes)
    if notes:
        return notes
    synthesized: list[LegacyPhenotypeNote] = []
    for note in report.patient_phenotypes or []:
        synthesized.append(
            LegacyPhenotypeNote(
                hpo_date=note.hpo_date or report.test_info.order,
                hpo_terms=note.hpo_terms,
                parsed_terms=note.parsed_terms,
                source_repeat_instance=None,
                source_row=None,
            )
        )
    return synthesized


def synthesize_linked_diagnosis_groups(
    report: LinkedGeneticReport,
) -> list[GeneticDiagnosisGroup]:
    groups = list(report.linked_diagnosis_groups)
    if groups:
        return groups
    diagnoses = [
        diagnosis_with_report_defaults(diagnosis, report)
        for diagnosis in (report.diagnoses or [])
    ]
    if not diagnoses and not report.diagnoses_note:
        return []
    return [
        GeneticDiagnosisGroup(diagnoses=chunk, diagnoses_note=report.diagnoses_note)
        for chunk in chunk_diagnoses(diagnoses, size=3)
    ]


def export_cases_to_redcap_rows(
    cases: list[LinkedGeneticsCase],
    *,
    header: Optional[list[str]] = None,
    include_unrelated_rows: bool = True,
    emit_base_row_when_missing: bool = False,
) -> tuple[list[str], list[dict[str, str]]]:
    if not header:
        header = None
        for case in cases:
            if case.redcap_column_order:
                header = list(case.redcap_column_order)
                break
        if not header:
            header = GENETICS_ONLY_HEADER.copy()

    all_rows: list[dict[str, str]] = []
    for case in cases:
        used_instances: dict[str, set[int]] = defaultdict(set)
        generated_by_instrument: dict[str, list[tuple[int, dict[str, str]]]] = (
            defaultdict(list)
        )

        # genetic_tests rows
        for report in case.reports:
            preferred_instance = report.source_repeat_instance
            instance = choose_instance(
                preferred_instance, used_instances["genetic_tests"]
            )
            row = empty_row(header)
            set_if_present(row, "redcap_repeat_instrument", "genetic_tests")
            set_if_present(row, "redcap_repeat_instance", instance)
            if report.source_row:
                for key, value in report.source_row.items():
                    if key in row:
                        row[key] = "" if value is None else str(value)
            set_if_present(row, "mrn_id", case.patient.mrn_id)
            set_if_present(row, "redcap_repeat_instrument", "genetic_tests")
            set_if_present(row, "redcap_repeat_instance", instance)
            apply_test_report_to_row(row, report)
            generated_by_instrument["genetic_tests"].append((instance, row))

            for note in synthesize_linked_phenotype_rows(report):
                preferred_phenotype_instance = note.source_repeat_instance
                phenotype_instance = choose_instance(
                    preferred_phenotype_instance, used_instances["patient_phenotypes"]
                )
                phenotype_row = empty_row(header)
                set_if_present(phenotype_row, "mrn_id", case.patient.mrn_id)
                set_if_present(
                    phenotype_row, "redcap_repeat_instrument", "patient_phenotypes"
                )
                set_if_present(
                    phenotype_row, "redcap_repeat_instance", phenotype_instance
                )
                if note.source_row:
                    for key, value in note.source_row.items():
                        if key in phenotype_row:
                            phenotype_row[key] = "" if value is None else str(value)
                apply_phenotype_to_row(
                    phenotype_row, note, fallback_order_date=report.test_info.order
                )
                generated_by_instrument["patient_phenotypes"].append(
                    (phenotype_instance, phenotype_row)
                )

            for group in synthesize_linked_diagnosis_groups(report):
                diagnoses = [
                    diagnosis_with_report_defaults(dx, report) for dx in group.diagnoses
                ]
                group_chunks = chunk_diagnoses(diagnoses, size=3) if diagnoses else [[]]
                preferred_group_instance = group.source_repeat_instance
                for chunk_index, chunk in enumerate(group_chunks):
                    preferred_instance = (
                        preferred_group_instance if chunk_index == 0 else None
                    )
                    diagnosis_instance = choose_instance(
                        preferred_instance, used_instances["genetic_diagnoses"]
                    )
                    diagnosis_row = empty_row(header)
                    set_if_present(diagnosis_row, "mrn_id", case.patient.mrn_id)
                    set_if_present(
                        diagnosis_row, "redcap_repeat_instrument", "genetic_diagnoses"
                    )
                    set_if_present(
                        diagnosis_row, "redcap_repeat_instance", diagnosis_instance
                    )
                    if group.source_row and chunk_index == 0:
                        for key, value in group.source_row.items():
                            if key in diagnosis_row:
                                diagnosis_row[key] = "" if value is None else str(value)
                    apply_diagnosis_group_to_row(
                        diagnosis_row, chunk, group.diagnoses_note
                    )
                    generated_by_instrument["genetic_diagnoses"].append(
                        (diagnosis_instance, diagnosis_row)
                    )

        # Unlinked phenotype rows
        for note in case.unlinked_patient_phenotypes:
            instance = choose_instance(
                note.source_repeat_instance, used_instances["patient_phenotypes"]
            )
            row = empty_row(header)
            set_if_present(row, "mrn_id", case.patient.mrn_id)
            set_if_present(row, "redcap_repeat_instrument", "patient_phenotypes")
            set_if_present(row, "redcap_repeat_instance", instance)
            if note.source_row:
                for key, value in note.source_row.items():
                    if key in row:
                        row[key] = "" if value is None else str(value)
            apply_phenotype_to_row(row, note)
            generated_by_instrument["patient_phenotypes"].append((instance, row))

        # Unlinked diagnosis groups
        for group in case.unlinked_diagnosis_groups:
            chunks = (
                chunk_diagnoses(group.diagnoses, size=3) if group.diagnoses else [[]]
            )
            for chunk_index, chunk in enumerate(chunks):
                preferred_instance = (
                    group.source_repeat_instance if chunk_index == 0 else None
                )
                instance = choose_instance(
                    preferred_instance, used_instances["genetic_diagnoses"]
                )
                row = empty_row(header)
                set_if_present(row, "mrn_id", case.patient.mrn_id)
                set_if_present(row, "redcap_repeat_instrument", "genetic_diagnoses")
                set_if_present(row, "redcap_repeat_instance", instance)
                if group.source_row and chunk_index == 0:
                    for key, value in group.source_row.items():
                        if key in row:
                            row[key] = "" if value is None else str(value)
                apply_diagnosis_group_to_row(row, chunk, group.diagnoses_note)
                generated_by_instrument["genetic_diagnoses"].append((instance, row))

        # Base row
        if case.legacy_base_row is not None or emit_base_row_when_missing:
            base_row = empty_row(header)
            if case.legacy_base_row:
                for key, value in case.legacy_base_row.items():
                    if key in base_row:
                        base_row[key] = "" if value is None else str(value)
            populate_patient_base_fields(base_row, case.patient)
            set_if_present(base_row, "redcap_repeat_instrument", "")
            set_if_present(base_row, "redcap_repeat_instance", "")
            all_rows.append(base_row)

        if include_unrelated_rows:
            legacy_by_form: dict[str, list[LegacyRepeatRow]] = defaultdict(list)
            for row_obj in case.legacy_other_repeat_rows:
                legacy_by_form[row_obj.redcap_repeat_instrument or ""].append(row_obj)
            for instrument in FORM_ORDER[1:]:
                if instrument in {
                    "genetic_tests",
                    "patient_phenotypes",
                    "genetic_diagnoses",
                }:
                    for _instance, row in sorted(
                        generated_by_instrument.get(instrument, []),
                        key=lambda item: item[0],
                    ):
                        all_rows.append(
                            {
                                column: blank_or_none(row.get(column))
                                for column in header
                            }
                        )
                for row_obj in sorted(
                    legacy_by_form.get(instrument, []),
                    key=lambda obj: (
                        obj.redcap_repeat_instance or 0,
                        obj.source_index or 0,
                    ),
                ):
                    preserved = empty_row(header)
                    for key, value in row_obj.row_data.items():
                        if key in preserved:
                            preserved[key] = "" if value is None else str(value)
                    all_rows.append(preserved)
            # Any generated instruments outside FORM_ORDER
            known_forms = set(FORM_ORDER)
            for instrument, rows_for_instrument in generated_by_instrument.items():
                if instrument in known_forms:
                    continue
                for _instance, row in sorted(
                    rows_for_instrument, key=lambda item: item[0]
                ):
                    all_rows.append(
                        {column: blank_or_none(row.get(column)) for column in header}
                    )
        else:
            for instrument in [
                "genetic_tests",
                "patient_phenotypes",
                "genetic_diagnoses",
            ]:
                for _instance, row in sorted(
                    generated_by_instrument.get(instrument, []),
                    key=lambda item: item[0],
                ):
                    all_rows.append(
                        {column: blank_or_none(row.get(column)) for column in header}
                    )

    return header, all_rows


# ---------------------------------------------------------------------------
# JSON coercion helpers for export-redcap
# ---------------------------------------------------------------------------


def report_to_case(
    report: GeneticReportExtraction, *, source_file: Optional[str] = None
) -> LinkedGeneticsCase:
    linked_report = LinkedGeneticReport.model_validate(report.model_dump(mode="json"))
    return LinkedGeneticsCase(
        patient=report.patient,
        nicu_flags=None,
        reports=[linked_report],
        unlinked_patient_phenotypes=[],
        unlinked_diagnosis_groups=[],
        legacy_base_row=None,
        legacy_other_repeat_rows=[],
        redcap_column_order=None,
        source_file=source_file,
    )


def load_json_payload(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def coerce_cases_from_json_payload(
    payload: Any, *, source_file: Optional[str] = None
) -> list[LinkedGeneticsCase]:
    if isinstance(payload, list):
        cases: list[LinkedGeneticsCase] = []
        for item in payload:
            cases.extend(coerce_cases_from_json_payload(item, source_file=source_file))
        return cases

    if not isinstance(payload, dict):
        raise ValueError("JSON input must be an object or a list of objects")

    # Full linked case JSON
    if (
        "reports" in payload
        or "legacy_base_row" in payload
        or "unlinked_diagnosis_groups" in payload
    ):
        return [get_case_model().model_validate(payload)]

    # Single report extraction JSON
    if "test_info" in payload and "interpretation" in payload:
        report = get_extraction_model().model_validate(payload)
        return [report_to_case(report, source_file=source_file)]

    raise ValueError(
        "Could not recognize JSON schema. Expected a linked case object/list or a single extracted report object."
    )


# ---------------------------------------------------------------------------
# CLI helpers and task runners
# ---------------------------------------------------------------------------


def normalize_base_url(base_url: str) -> str:
    base = base_url.rstrip("/")
    if base.endswith("/v1"):
        return base
    return f"{base}/v1"


def choose_backend(args: argparse.Namespace, model_id: str) -> ChatBackend:
    backend_name = args.backend
    if backend_name == "auto":
        backend_name = "vllm" if args.base_url else "hf"
    if backend_name == "vllm":
        if not args.base_url:
            raise SystemExit("--backend vllm requires --base-url")
        return VLLMBackend(
            model_id,
            base_url=args.base_url,
            read_timeout=args.read_timeout,
            api_key=args.api_key,
        )
    return HFBackend(model_id, num_gpus=args.num_gpus, dtype=args.dtype)


def default_extract_output_path(input_path: Path) -> Path:
    return input_path.with_suffix(".extracted.json")


def default_import_output_path(input_path: Path) -> Path:
    return input_path.with_suffix(".linked.json")


def default_export_output_path(input_path: Path) -> Path:
    return input_path.with_suffix(".redcap.csv")


def default_columns_output_path(data_dictionary_path: Path) -> Path:
    return data_dictionary_path.with_suffix(".redcap_columns.json")


def maybe_assign_source_file(
    report: GeneticReportExtraction, input_path: Path
) -> GeneticReportExtraction:
    if report.test_info.file:
        return report
    updated_test_info = report.test_info.model_copy(
        update={"file": str(input_path.name)}
    )
    return report.model_copy(update={"test_info": updated_test_info})


def run_extract(args: argparse.Namespace) -> None:
    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")
    if not args.model:
        raise SystemExit("extract task requires --model")

    report_text = load_report(args.input, args.input_format)
    logger.info("loaded %d characters from %s", len(report_text), args.input)

    model_id = resolve_model(args.model)
    logger.info("model resolved: %s -> %s", args.model, model_id)

    model_cls = get_extraction_model()
    prompt_messages = build_messages(report_text, model_cls)
    if args.print_prompt:
        print(json.dumps(prompt_messages, indent=2, ensure_ascii=False))
        return

    backend = choose_backend(args, model_id)
    result = extract(
        report_text,
        backend,
        model_cls=model_cls,
        temperature=args.temperature,
        top_p=args.top_p,
        max_new_tokens=args.max_new_tokens,
        max_retries=args.max_retries,
    )
    if not isinstance(result, GeneticReportExtraction):
        result = GeneticReportExtraction.model_validate(result.model_dump(mode="json"))

    result = maybe_assign_source_file(result, args.input)
    output_data = result.model_dump(mode="json")
    output_path = args.output or default_extract_output_path(args.input)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(output_data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    logger.info("wrote %s", output_path)
    print(json.dumps(output_data, indent=2, ensure_ascii=False))

    if args.redcap_output:
        cases = [report_to_case(result, source_file=str(args.input))]
        if args.data_dictionary:
            header = build_redcap_column_order_from_dictionary(args.data_dictionary)
            include_unrelated = True
            emit_base_row = args.emit_base_row
        else:
            header = GENETICS_ONLY_HEADER.copy()
            include_unrelated = False
            emit_base_row = args.emit_base_row
        redcap_header, redcap_rows = export_cases_to_redcap_rows(
            cases,
            header=header,
            include_unrelated_rows=include_unrelated,
            emit_base_row_when_missing=emit_base_row,
        )
        args.redcap_output.parent.mkdir(parents=True, exist_ok=True)
        write_rows_to_delimited(args.redcap_output, redcap_header, redcap_rows)
        logger.info("wrote REDCap-compatible rows to %s", args.redcap_output)


def run_import_redcap(args: argparse.Namespace) -> None:
    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")

    header, rows = read_tabular_rows(args.input)
    logger.info("loaded %d rows from %s", len(rows), args.input)
    cases = import_redcap_rows_to_cases(header, rows, source_file=str(args.input))
    output_path = args.output or default_import_output_path(args.input)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_data = [case.model_dump(mode="json") for case in cases]
    output_path.write_text(
        json.dumps(output_data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    logger.info("wrote %s", output_path)
    print(json.dumps(output_data, indent=2, ensure_ascii=False))


def run_export_redcap(args: argparse.Namespace) -> None:
    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")

    payload = load_json_payload(args.input)
    cases = coerce_cases_from_json_payload(payload, source_file=str(args.input))

    header: Optional[list[str]] = None
    include_unrelated = not args.genetics_only
    if args.data_dictionary:
        header = build_redcap_column_order_from_dictionary(args.data_dictionary)
    elif args.genetics_only:
        header = GENETICS_ONLY_HEADER.copy()
    else:
        for case in cases:
            if case.redcap_column_order:
                header = list(case.redcap_column_order)
                break
        if header is None:
            header = GENETICS_ONLY_HEADER.copy()
            include_unrelated = False

    export_header, export_rows = export_cases_to_redcap_rows(
        cases,
        header=header,
        include_unrelated_rows=include_unrelated,
        emit_base_row_when_missing=args.emit_base_row,
    )
    output_path = args.output or default_export_output_path(args.input)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_rows_to_delimited(output_path, export_header, export_rows)
    logger.info("wrote %s", output_path)
    print(
        json.dumps(
            {"rows_written": len(export_rows), "output": str(output_path)}, indent=2
        )
    )


def run_print_redcap_columns(args: argparse.Namespace) -> None:
    if not args.data_dictionary:
        raise SystemExit("print-redcap-columns requires --data-dictionary")
    header = build_redcap_column_order_from_dictionary(args.data_dictionary)
    output_path = args.output or default_columns_output_path(args.data_dictionary)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(header, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    logger.info("wrote %s", output_path)
    print(json.dumps(header, indent=2, ensure_ascii=False))


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pediatric/NICU genetic report extractor with REDCap-compatible import/export",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Model aliases: "
            + ", ".join(
                f"{alias} -> {model_id}" for alias, model_id in MODEL_ALIASES.items()
            )
        ),
    )
    parser.add_argument(
        "--task",
        choices=["extract", "import-redcap", "export-redcap", "print-redcap-columns"],
        default="extract",
        help="Primary task (default: extract)",
    )
    parser.add_argument("-i", "--input", type=Path, help="Input file path")
    parser.add_argument(
        "-o", "--output", type=Path, default=None, help="Output file path"
    )
    parser.add_argument(
        "-m",
        "--model",
        default=None,
        help=(
            "Model alias or full model ID / served model name for extract task. "
            f"Known aliases: {', '.join(SUPPORTED_ALIASES)}"
        ),
    )
    parser.add_argument(
        "--backend",
        choices=["auto", "hf", "vllm"],
        default="auto",
        help="Inference backend for extract (default: auto)",
    )
    parser.add_argument(
        "--base-url",
        default=None,
        help="Base URL for an OpenAI-compatible vLLM server, e.g. http://127.0.0.1:8020 or http://127.0.0.1:8020/v1",
    )
    parser.add_argument(
        "--api-key",
        default=os.environ.get("VLLM_API_KEY") or os.environ.get("OPENAI_API_KEY"),
        help="Optional API key for the endpoint backend",
    )
    parser.add_argument(
        "--read-timeout",
        type=int,
        default=1800,
        help="Endpoint read timeout in seconds (extract/vLLM only)",
    )
    parser.add_argument(
        "--num-gpus",
        type=int,
        default=1,
        help="HF backend only: number of visible GPUs to budget across",
    )
    parser.add_argument(
        "--dtype",
        choices=["auto", "bfloat16", "float16", "float32"],
        default="auto",
        help="HF backend only: model dtype",
    )
    parser.add_argument(
        "--input-format",
        choices=["text", "pdf"],
        default=None,
        help="Override automatic report input-format detection (extract only)",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=0.0,
        help="Sampling temperature for extract (0 for deterministic decoding)",
    )
    parser.add_argument(
        "--top-p",
        type=float,
        default=1.0,
        help="Nucleus sampling parameter when temperature > 0 (extract only)",
    )
    parser.add_argument(
        "--max-new-tokens",
        type=int,
        default=32768,
        help="Maximum tokens to generate (extract only)",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=3,
        help="Validation-aware retry attempts (extract only)",
    )
    parser.add_argument(
        "--print-prompt",
        action="store_true",
        help="Print the schema-derived prompt and exit (extract only)",
    )
    parser.add_argument(
        "--data-dictionary",
        type=Path,
        default=None,
        help="Optional REDCap data dictionary .xlsx used to derive full column order",
    )
    parser.add_argument(
        "--redcap-output",
        type=Path,
        default=None,
        help="Optional extra CSV/TSV export written during extract",
    )
    parser.add_argument(
        "--genetics-only",
        action="store_true",
        help="For export-redcap, emit only genetics-related columns/rows instead of trying to preserve the full raw export layout",
    )
    parser.add_argument(
        "--emit-base-row",
        action="store_true",
        help="Emit a non-repeating base row even when exporting from a stand-alone extracted report",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def main(argv: Optional[list[str]] = None) -> None:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    if args.task == "print-redcap-columns":
        run_print_redcap_columns(args)
        return

    if args.input is None:
        raise SystemExit(f"--input is required for task '{args.task}'")

    if args.task == "extract":
        run_extract(args)
    elif args.task == "import-redcap":
        run_import_redcap(args)
    elif args.task == "export-redcap":
        run_export_redcap(args)
    else:  # pragma: no cover - parser already constrains choices
        raise SystemExit(f"Unsupported task: {args.task}")


if __name__ == "__main__":
    main()
